import tkinter as tk
from tkinter import messagebox, ttk
import logging
from datetime import date, datetime, timedelta
from dateutil.relativedelta import relativedelta
from dateutil import parser
from openpyxl import load_workbook
import xlwings as xw
import os, sys
import shutil
import re

from helper.toggl_parse_data import get_toggl_time_entries
   
def update_entries_in_anw (time_entry_list, folder, file, file_new, workingtime_by_day_list):
    logging.info("update entries in anw")
    wb = load_workbook(folder + file)
    anw = wb["ANW"]
    
    # Enter working hours
    for project in time_entry_list:
        logging.debug("project: " + project)
        
        switch_col_number = -1
        for col in range(8,100):
            if anw.cell(row=4, column=col).value == "angerechn. Reisezeit":
                switch_col_number = col
                break
            
        if switch_col_number == -1:
            logging.error("Column ""angerechn. Reisezeit"" not found in row 4 of the ANW")
            raise KeyError("Column ""angerechn. Reisezeit"" not found in row 4 of the ANW")
        
        if "reisen" not in project.lower():
            col_start = 8
            col_end = switch_col_number - 1
        else:
            col_start = switch_col_number + 1
            col_end = 43
        
        project_col = -1
        excel_proj = "empty"
        
        for col in range(col_start, col_end):
            if anw.cell(row=4, column=col).value is None or anw.cell(row=4, column=col).value == "":
                continue
            excel_proj = anw.cell(row=4, column=col).value[:4] # type: ignore
            if excel_proj == project[:4]:
                project_col = col
                break
        
        if project_col == -1 or excel_proj != project[:4]:
            logging.error("project '" + str(project) + "' not found in excel")
            raise KeyError("project '" + str(project) + "' not found in excel")
        
        
        for date_str in time_entry_list[project]:
            logging.debug("date: " + date_str)
            logging.debug("hours: " + str(time_entry_list[project][date_str]["hours"]))
            
            row_offset = 5
            
            day =  parser.parse(date_str).day
            
            anw.cell(row = day + row_offset, column=project_col).value = time_entry_list[project][date_str]["hours"]
    
    # Working times per day       
    for date_str in workingtime_by_day_list:
        logging.debug("date: " + str(date_str))
        logging.debug("hours: " + str(workingtime_by_day_list[date_str]))
        
        row_offset = 5
        
        day = parser.parse(date_str).day
        if  workingtime_by_day_list[date_str]["endtime"].hour == 0:
            hour = 24
        else:
            hour = int(workingtime_by_day_list[date_str]["endtime"].strftime("%H"))
        
        anw.cell(row = day + row_offset, column=3).value = workingtime_by_day_list[date_str]["starttime"].hour + workingtime_by_day_list[date_str]["starttime"].minute/100
        anw.cell(row = day + row_offset, column=4).value = hour + workingtime_by_day_list[date_str]["endtime"].minute/100
                    
    wb.save(folder + file_new)

def update_entries_in_anw_new (time_entry_list, folder, file, workingtime_by_day_list):
    logging.debug("In update_entries_in_anw_new")
    wb = xw.Book(file)
    logging.debug("End of update_entries_in_anw_new")

# The function to update the cell M1, save the new file, and delete the old one
def adjust_anw_for_new_month(folder, file, file_new):
    wb = load_workbook(folder + file, data_only=True)
    
    # Updating M1
    m1_value = wb["ANW"]["M1"].value
    if isinstance(m1_value, datetime):
        new_value = m1_value + relativedelta(months=1)
        wb["ANW"]["M1"].value = new_value
    else:
        raise ValueError("Cell M1 does not contain a date.")
    
    wb.save(folder + file_new)
    
    os.remove(folder + file)

    logging.info(f"New file saved: {folder + file_new}")
    logging.info(f"Old file deleted: {folder + file}")
    
def increment_month_in_filename(filename):
    match = re.search(r'(\d{6})\.xlsx$', filename)
    if match:
        date_str = match.group(1)
        date_obj = datetime.strptime(date_str, "%Y%m")
        new_date_obj = date_obj + relativedelta(months=1)
        new_date_str = new_date_obj.strftime("%Y%m")
        new_filename = filename.replace(date_str, new_date_str)
        return new_filename
    else:
        raise ValueError("Filename does not contain a valid date format")

# Popup function
def ask_user_confirmation(folder, file_template, file_with_workinghours):
    root = tk.Tk()
    root.withdraw()  # Hides the main window
    
    file_template_new_month = increment_month_in_filename(file_template)
    
    result = messagebox.askyesno("Confirmation", "Do you want to adjust the ANW template for the next month? " + file_template + " will become " 
                                 + file_template_new_month + ".")
    
    if result:
        logging.info("User confirmed to adjust template for new month")
        adjust_anw_for_new_month(folder, file, file_template_new_month)
        
        result2 = messagebox.askyesno("Confirmation", "Do you want the filled ANW with workinghours to be renamed from " + file_with_workinghours + 
                                      " to " + file + "?")
    
        if  result2:
            logging.info("User confirmed to rename the file with working hours")
            os.rename(folder + file_with_workinghours, folder + file)
            
        else:
            logging.info("User declined to rename the file with working hours")
            
    else:
        logging.info("User declined to adjust template for new month")
    
    root.destroy()

# Function to ask the user to select an Excel file using a list box
def ask_user_for_file(files):
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    
    def on_select(event):
        selected_file = listbox.get(listbox.curselection())
        root.selected_file = selected_file

    def on_close():
        if hasattr(root, 'selected_file'):
            root.destroy()  # Close the window
        else:
            messagebox.showerror("Error", "No file selected. Please try again.")
    
    top = tk.Toplevel(root)
    top.title("Select an Excel file")

    tk.Label(top, text="Multiple Excel files found:").pack(pady=5)
    
    listbox = tk.Listbox(top)
    listbox.pack(padx=10, pady=5)
    for file in files:
        listbox.insert(tk.END, file)
    listbox.bind('<<ListboxSelect>>', on_select)
        
    # Button to close window if no selection made
    button = tk.Button(top, text="Close", command=on_close)
    button.pack(pady=5)
    
    # root.wait_window(top)
    root.mainloop()
    
    return getattr(root, 'selected_file', None)

if __name__ == '__main__':

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        handlers=[
            # logging.FileHandler("debug.log"), # For logging to file
            logging.StreamHandler()
        ]
    )
    gettrace = getattr(sys, 'gettrace', None)
    if gettrace():
        logging.getLogger().setLevel(logging.DEBUG)

    logging.info("----------------------------------------")
    logging.info("Starting Toggl to Jira Transfer")
    logging.debug("Debugging is enabled")
    
    linux_folder = "/mnt/c/Users/PrV/OneDrive - viadee Unternehmensberatung AG/Arbeitsnachweis/"
    windows_folder = linux_folder.replace('/mnt/c', 'C:').replace('/', '\\')
    if sys.platform == 'linux': # execution in WSL Linux
        folder = linux_folder
    else: # execution probably in windows
        folder = windows_folder
    
    excel_files = [f for f in os.listdir(folder) if f.endswith('.xlsx')]

    # Check if there are multiple Excel files
    if len(excel_files) > 1:
        file = ask_user_for_file(excel_files)
        if not file:
            raise FileNotFoundError("User did not select a valid file.")
    elif len(excel_files) == 1:
        file = excel_files[0]
    else:
        raise FileNotFoundError(f"No Excel files found in folder: {folder}")
        
    match = re.search(r'Anw_PrV_(\d{6}).*\.xlsx$', file)
    matchOrig = re.search(r'Anw_PrV_(\d{6})\.xlsx$', file)
    if match:
        fileshort = match.group(1)
        start_date = datetime.strptime(fileshort, "%Y%m").date().replace(day=1)
    else:
        raise ValueError(f"Filename '{file}' does not match the expected format 'Anw_PrV_YYYYMM*.xlsx'")
        
    end_date = (start_date + timedelta(days=32)).replace(day=1) # start of next month
    logging.debug("Start Date: " + str(start_date))
    logging.debug("End Date: " + str(end_date))
        
    time_entry_list, workingtime_by_day_list, time_entry_list_detail = get_toggl_time_entries(start_date, end_date)

    # time_entry_list = pickle.load(open("time_entry_list.pickle", "rb"))
    if matchOrig:
        file_new = file.replace(".xlsx", "n.xlsx")
        # Copy the file before making updates
        logging.info(f"Copying original Excel '{file}' to '{file_new}'. Existing Files will be replaced.")
        shutil.copy(folder + file, folder + file_new)
        
    # update_entries_in_anw(time_entry_list, folder, file, file, workingtime_by_day_list)
    update_entries_in_anw_new(time_entry_list, folder, file, workingtime_by_day_list)
    
    logging.info("Finished Toggl to Anw Transfer")
    
    # open created file
    logging.info("Opening new filled ANW file " + file)
    os.system(f'explorer.exe "{windows_folder + file}"')
    
    # ask_user_confirmation(folder, file, file_with_workinghours) # formula get lost currently -> not usable. Maybe use xlwings library alternatively